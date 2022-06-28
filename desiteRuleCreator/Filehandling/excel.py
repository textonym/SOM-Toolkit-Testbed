import os.path

import openpyxl
import logging
from desiteRuleCreator.QtDesigns import ui_mainwindow
from desiteRuleCreator.data import classes, constants
from desiteRuleCreator.Widgets import object_widget
from desiteRuleCreator import logs

def transform_value_types(value: str):
    special = False
    if value is not None:
        if value.lower() in ["string", ]:
            data_type = constants.XS_STRING
        elif value.lower() in ["double"]:
            data_type = constants.XS_DOUBLE
        elif value.lower() in ["boolean", "bool"]:
            data_type = constants.XS_BOOL
        elif value.lower() in ["int", "integer"]:
            data_type = constants.XS_INT
        else:
            special = True
            data_type = constants.XS_STRING
    else:
        data_type = constants.XS_STRING

    return data_type, special

def split_string(text: str):
    if text is None:
        return None
    text = text.split(";")
    for i, item in enumerate(text):
        if "(" in item:
            item = item.split("(")
            text[i] = item[0]
        text[i] = text[i].strip()

    return text

def link_psets(pset, cell, pset_dict, sheet, obj=None, debug=False):
    # pset_dict[kuerzel] = [pset, cell]



    if debug:
        print(pset.name)

    elternklasse = sheet.cell(cell.row + 2, cell.column + 1).value
    elternklassen: list = split_string(elternklasse)

    for elternklasse in elternklassen:
        if elternklasse != "AE" and elternklasse != "-":
            value = pset_dict.get(elternklasse.upper())
            if value is not None:
                [eltern_pset, eltern_cell,dummy] = value

                if obj is not None:
                    new_pset = classes.PropertySet(eltern_pset.name)
                    eltern_pset.add_child(new_pset)
                    obj.add_property_set(new_pset)
                link_psets(eltern_pset, eltern_cell, pset_dict, sheet, obj, debug=debug)
            else:
                logging.warning(f"ACHTUNG {sheet.cell(cell.row, cell.column + 1).value} hat einen Fehler in der Elternklasse ({elternklasse.upper()} existiert nicht!)")


def iterate_entries(pset, sheet, entry, cell_list):
    special_values = list()
    while entry.value is not None and entry not in cell_list:
        data_type, special = transform_value_types(sheet.cell(row=entry.row, column=entry.column + 2).value)
        classes.Attribute(pset, entry.value, "", constants.VALUE_TYPE_LOOKUP[constants.LIST], data_type=data_type)
        if special:
            special_values.append(entry)

        entry = sheet.cell(row=entry.row + 1, column=entry.column)
    return special_values


def create_predefined_pset(sheet, cell, cell_list):
    name = sheet.cell(row=cell.row, column=cell.column + 1).value
    kuerzel = sheet.cell(row=cell.row + 1, column=cell.column + 1).value.upper()
    elternklasse = sheet.cell(row=cell.row + 2, column=cell.column + 1).value

    pset = classes.PropertySet(name)

    entry = sheet.cell(row=cell.row + 4, column=cell.column)
    special_values = iterate_entries(pset, sheet, entry, cell_list)

    return pset, kuerzel, special_values


def create_object(sheet, cell, pset_dict, cell_list):
    name = sheet.cell(row=cell.row, column=cell.column + 1).value
    kuerzel = sheet.cell(row=cell.row + 1, column=cell.column + 1).value.upper()
    aggregate_children = sheet.cell(row=cell.row + 3, column=cell.column + 1).value
    ident = sheet.cell(row=cell.row, column=cell.column + 2).value

    pset = classes.PropertySet(name)

    entry = sheet.cell(row=cell.row + 5, column=cell.column)
    special_values = iterate_entries(pset, sheet, entry, cell_list)

    ident_pset = classes.PropertySet("Allgemeine Eigenschaften")
    parent: classes.PropertySet = pset_dict["AE"][0]
    parent.add_child(ident_pset)
    ident_attrib: classes.Attribute = ident_pset.get_attribute_by_name("bauteilKlassifikation")

    ident_attrib.value = [ident]
    obj = classes.Object(name, ident_attrib)
    obj.add_property_set(ident_pset)
    obj.add_property_set(pset)

    aggregate_list = split_string(aggregate_children)
    if aggregate_list is None:
        logging.warning(f"Achtung! {name} besitzt keinen Wert bei 'Besteht aus'")
        aggregate_list=[]
    elif  aggregate_list == ["-"]:
        aggregate_list = []

    return obj, special_values, pset, kuerzel,aggregate_list


def start(main_window, path):
    base_path = os.path.dirname(logs.__file__)
    base_path = os.path.join(base_path,"log.txt")
    os.remove(base_path)
    logging.basicConfig(filename=base_path,level=logging.WARNING)
    logging.warning("----------------------------------------------")
    book = openpyxl.load_workbook(path)
    sheet = book.active

    row: (sheet.cell())
    name_cells = list()
    for row in sheet:
        for cell in row:
            value = cell.value
            if value is not None:
                text = value.strip()
                if text in ["name", "name:"]:
                    if sheet.cell(cell.row + 1, cell.column).value == "Kürzel":
                        name_cells.append(cell)
                    else:
                        logging.warning(f"{sheet.cell(cell.row + 1, cell.column)} hat den Wert 'name'")


    pset_dict = dict()

    special_values = list()
    aggregate_dict = dict()

    for cell in name_cells:
        ident_value = sheet.cell(row=cell.row, column=cell.column + 2).value

        if ident_value is None:
            pset, kuerzel, special = create_predefined_pset(sheet, cell, name_cells)
            pset_dict[kuerzel] = [pset, cell,None]
            special_values += special

        else:
            obj, special, pset, kuerzel,aggregate_list = create_object(sheet, cell, pset_dict, name_cells)
            aggregate_dict[obj] = aggregate_list
            if pset_dict.get(kuerzel) is not None:
                logging.warning(f"Kuerzel {kuerzel} für {obj.name} und {pset_dict[kuerzel][0].name} identisch!")
            pset_dict[kuerzel] = [pset, cell,obj]
            special_values += special

    for kuerzel, (pset, cell,obj) in pset_dict.items():
        link_psets(pset, cell, pset_dict, sheet, pset.object, debug=False)

    tree_dict = dict()

    obj: classes.Object
    for obj in classes.Object:
        item = object_widget.add_object_to_tree(main_window,obj,None)
        tree_dict[obj.ident_attrib.value[0]] = item

    ident: str
    for ident, item in tree_dict.items():
        ident_list = ident.split(".")
        ident_list = ident_list[:-1]
        ident = ".".join(ident_list)
        parent_item: classes.CustomTreeItem = tree_dict.get(ident)
        if parent_item is not None:
            root = item.treeWidget().invisibleRootItem()
            item = root.takeChild(root.indexOfChild(item))
            parent_item.addChild(item)

    for obj in classes.Object:
        aggregate_list = aggregate_dict[obj]
        for kuerzel in aggregate_list:
            dic = pset_dict.get(kuerzel)
            if dic is not None:
                obj_child = dic[2]
                if obj_child is not None:
                    obj.add_aggregation(obj_child)


                else:logging.warning(f"[{obj.name}] Aggregation: Kürzel {kuerzel} existiert nicht")
            else:logging.warning(f"[{obj.name}] Aggregation: Kürzel {kuerzel} existiert nicht")
